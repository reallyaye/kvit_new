import argparse
import os
import sys

from config import RECEIPTS_DIR
from database import (
    get_db,
    migrate_db,
    purge_missing_receipts,
    sync_receipts_with_filesystem,
    write_transaction,
)
from database.migrations import migrate_receipts_to_sharding
from services.pdf import pdf_processor
from services.security import validate_safe_path


def load_excel_rows(file_path: str):
    """Считывает строки из Excel файла (.xlsx через openpyxl или .xls через xlrd)."""
    ext = os.path.splitext(file_path)[1].lower()
    rows = []

    if ext == '.xlsx':
        try:
            import openpyxl
            book = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet = book[book.sheetnames[0]]
            for r in sheet.iter_rows(min_row=2, values_only=True):
                if any(r):
                    rows.append(list(r))
            book.close()
        except ImportError:
            print("❌ Ошибка: модуль 'openpyxl' не установлен. Установите: pip install openpyxl")
            sys.exit(1)
    elif ext == '.xls':
        try:
            import xlrd
            book = xlrd.open_workbook(file_path)
            sheet = book.sheet_by_index(0)
            for r_idx in range(1, sheet.nrows):
                row_vals = sheet.row_values(r_idx)
                rows.append(row_vals)
        except ImportError:
            print("❌ Ошибка: модуль 'xlrd' не установлен. Установите: pip install xlrd")
            sys.exit(1)
    else:
        print(f"❌ Ошибка: неподдерживаемый формат файла: {ext}. Поддерживаются .xls и .xlsx")
        sys.exit(1)

    return rows

def main():
    parser = argparse.ArgumentParser(description="Импорт реестра лицевых счетов и PDF квитанций в базу данных.")
    parser.add_argument('--accounts', help="Путь к Excel-файлу с лицевыми счетами (.xls или .xlsx)")
    parser.add_argument('--receipts', help="Путь к PDF-файлу или папке с PDF-файлами квитанций")
    parser.add_argument('--reset', action='store_true', help="Сбросить существующие таблицы перед импортом")
    parser.add_argument('--migrate-sharding', action='store_true', help="Мигрировать существующие квитанции в 2-уровневую структуру папок")
    parser.add_argument('--sync-fs', action='store_true', help="Безопасная синхронизация с диском (перевод отсутствующих в статус missing без удаления)")
    parser.add_argument('--purge-missing', action='store_true', help="Явная очистка: безвозвратно удалить из БД все записи со статусом missing")
    args = parser.parse_args()

    if args.migrate_sharding:
        print("📂 Запуск миграции файлов квитанций в шардированную структуру...")
        migrated_files, updated_db = migrate_receipts_to_sharding()
        print(f"✅ Миграция завершена: перемещено файлов: {migrated_files}, обновлено записей в БД: {updated_db}")
        if not args.accounts and not args.receipts and not args.sync_fs and not args.purge_missing:
            sys.exit(0)

    if args.sync_fs:
        print("🔄 Проверка и безопасная синхронизация БД с файловым хранилищем receipts/...")
        marked_missing, restored_ready, valid = sync_receipts_with_filesystem()
        print(f"✅ Синхронизация завершена: помечено missing: {marked_missing}, восстановлено: {restored_ready}, актуальных на диске: {valid}")
        if not args.accounts and not args.receipts and not args.purge_missing:
            sys.exit(0)

    if args.purge_missing:
        print("🗑️ Очистка записей со статусом 'missing' из базы данных...")
        purged = purge_missing_receipts()
        print(f"✅ Очистка завершена: удалено {purged} записей.")
        if not args.accounts and not args.receipts:
            sys.exit(0)

    if not args.accounts and not args.receipts:
        parser.print_help()
        sys.exit(0)

    os.makedirs(RECEIPTS_DIR, exist_ok=True)
    con = get_db()

    try:
        # Базовая инициализация таблиц, если их ещё нет
        con.executescript('''
            CREATE TABLE IF NOT EXISTS accounts(
                id INTEGER PRIMARY KEY,
                account_number TEXT NOT NULL UNIQUE,
                customer_name TEXT,
                address TEXT,
                street TEXT,
                building TEXT,
                corpus TEXT,
                district TEXT,
                organization TEXT
            );
            CREATE INDEX IF NOT EXISTS idx_accounts_account ON accounts(account_number);

            CREATE TABLE IF NOT EXISTS receipts(
                id INTEGER PRIMARY KEY,
                account_number TEXT NOT NULL,
                period TEXT NOT NULL,
                pdf_file TEXT NOT NULL,
                content_hash TEXT,
                access_token TEXT,
                UNIQUE(account_number, period)
            );
            CREATE INDEX IF NOT EXISTS idx_receipts_account_period ON receipts(account_number, period);
        ''')
        con.commit()
    finally:
        con.close()

    # Запуск миграций для создания токенов и хешей
    migrate_db()

    con = get_db()
    try:
        if args.reset:
            print("⚠ Внимание: Запрошен сброс данных (--reset). Очистка таблиц...")
            con.execute("DELETE FROM receipts")
            con.execute("DELETE FROM accounts")
            con.commit()
            print("База данных успешно очищена.\n")

        # 1. Импорт лицевых счетов из Excel
        if args.accounts:
            try:
                accounts_path = validate_safe_path(args.accounts)
            except ValueError as e:
                print(f"❌ Ошибка безопасности пути к счетам: {e}")
                sys.exit(1)

            if not os.path.isfile(accounts_path):
                print(f"❌ Файл счетов не найден: {accounts_path}")
                sys.exit(1)

            print(f" Чтение файла лицевых счетов: {accounts_path}...")
            raw_rows = load_excel_rows(accounts_path)
            accounts_to_insert = []

            for row in raw_rows:
                if not row or row[0] in ('', None):
                    continue
                v = row[0]
                if isinstance(v, float) and v.is_integer():
                    v = str(int(v))
                account = str(v).strip()
                if not account:
                    continue

                def get_val(idx, current_row=row):
                    if idx < len(current_row) and current_row[idx] is not None:
                        val = current_row[idx]
                        if isinstance(val, float) and val.is_integer():
                            return str(int(val))
                        return str(val).strip()
                    return ''

                cust_name = get_val(1)
                addr = get_val(2)
                street = get_val(3) if len(row) > 3 else addr
                building = get_val(4) if len(row) > 4 else ''
                corpus = get_val(5) if len(row) > 5 else ''
                district = get_val(6) if len(row) > 6 else ''
                org = get_val(7) if len(row) > 7 else ''

                accounts_to_insert.append((account, cust_name, addr, street, building, corpus, district, org))

            if accounts_to_insert:
                with write_transaction() as con_write:
                    con_write.executemany('''
                        INSERT OR REPLACE INTO accounts(
                            account_number, customer_name, address, street, building, corpus, district, organization
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    ''', accounts_to_insert)
                print(f"✅ Успешно импортировано/обновлено счетов: {len(accounts_to_insert)}")

        # 2. Импорт PDF квитанций
        if args.receipts:
            try:
                receipts_path = validate_safe_path(args.receipts)
            except ValueError as e:
                print(f"❌ Ошибка безопасности пути к квитанциям: {e}")
                sys.exit(1)

            session_hashes = set()

            pdf_files = []
            if os.path.isfile(receipts_path):
                pdf_files.append(receipts_path)
            elif os.path.isdir(receipts_path):
                for root, _, files in os.walk(receipts_path):
                    for f in files:
                        if f.lower().endswith('.pdf'):
                            pdf_files.append(os.path.join(root, f))
            else:
                print(f"❌ Путь к PDF не найден: {receipts_path}")
                sys.exit(1)

            print(f"📄 Обработка PDF ({len(pdf_files)} шт.)...")
            total_added = 0
            total_orphan = 0
            total_skipped = 0
            total_duplicates = 0
            all_receipts = []

            for pdf_path in pdf_files:
                base_name = os.path.basename(pdf_path)
                added, orphan, skipped, dups, details, receipts = pdf_processor.process_single_pdf(
                    pdf_path, base_name, known_accounts=None, existing_hashes=session_hashes
                )
                total_added += added
                total_orphan += orphan
                total_skipped += skipped
                total_duplicates += dups
                all_receipts.extend(receipts)

            print(f"✅ Квитанций привязано к счетам: {total_added}")
            if total_orphan > 0:
                print(f"⚠  Квитанций без счёта в базе (сироты): {total_orphan}")
            if total_duplicates > 0:
                print(f"🔄 Пропущено дубликатов: {total_duplicates}")
            if total_skipped > 0:
                print(f"❌ Пропущено (не распознано): {total_skipped}")

        # 3. Отчет о сверке
        total_accounts = con.execute('SELECT COUNT(*) FROM accounts').fetchone()[0]
        total_receipts = con.execute('SELECT COUNT(*) FROM receipts').fetchone()[0]
        matched = con.execute('''
            SELECT COUNT(DISTINCT a.account_number)
            FROM accounts a
            JOIN receipts r ON r.account_number = a.account_number
        ''').fetchone()[0]
        unmatched_count = total_accounts - matched
        orphans = con.execute('''
            SELECT COUNT(*) FROM receipts r
            LEFT JOIN accounts a ON a.account_number = r.account_number
            WHERE a.id IS NULL
        ''').fetchone()[0]

        print()
        print('=' * 60)
        print('  ОТЧЁТ О СВЕРКЕ')
        print('=' * 60)
        print(f'  Лицевых счетов в базе:       {total_accounts}')
        print(f'  Квитанций загружено:          {total_receipts}')
        print(f'  Счетов С квитанцией:          {matched}')
        print(f'  Счетов БЕЗ квитанции:         {unmatched_count}')
        print(f'  Квитанций-сирот (нет счёта):  {orphans}')
        if total_accounts > 0:
            coverage = round(matched / total_accounts * 100, 1)
            print(f'  Процент покрытия:             {coverage}%')
        print('=' * 60)

    finally:
        con.close()

if __name__ == '__main__':
    main()

