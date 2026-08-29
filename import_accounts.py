#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Универсальный инструмент импорта, экспорта и миграции базы лицевых счетов.

Поддерживаемые форматы:
  - CSV / TSV / TXT (автоопределение кодировок UTF-8, CP1251 и разделителей ;, \t, |)
  - Excel (.xlsx, .xls)
  - JSON / JSONL
  - SQLite (.sqlite3, .db) — перенос данных напрямую из SQLite в PostgreSQL

Использование:
  python import_accounts.py --file accounts.csv
  python import_accounts.py --file old_data.sqlite3
  python import_accounts.py --export backup_accounts.csv
  python import_accounts.py --file registry.xlsx --mode replace
"""

import argparse
import csv
import io
import json
import os
import re
import sqlite3
import sys
import time
from typing import Any, Dict, Generator, List, Optional, Tuple

import config
from database.connection import get_db, is_postgres_configured, write_transaction
from database.migrations import migrate_db
from logger import logger

# Синонимы названий колонок для автоматического маппинга
COLUMN_SYNONYMS = {
    'account_number': [
        'account_number', 'account', 'account_no', 'acc', 'acc_num',
        'лицевой счет', 'лицевой_счет', 'лицевой_счёт', 'лицевой', 'лс', 'л/с',
        'номер счета', 'номер_счета', 'номер_счёта', 'счет', 'счёт', 'код',
        'жеке шот', 'жеке_шот', 'шот'
    ],
    'customer_name': [
        'customer_name', 'name', 'fio', 'full_name', 'client', 'payer',
        'фио', 'ф.и.о.', 'абонент', 'плательщик', 'потребитель', 'клиент',
        'собственник', 'жилец', 'имя', 'аты-жөні', 'тұтынушы'
    ],
    'address': [
        'address', 'addr', 'full_address', 'street_address',
        'адрес', 'полный адрес', 'мекенжай', 'мекенжайы', 'адрес проживания'
    ],
    'street': [
        'street', 'street_name', 'улица', 'көше', 'көшесі', 'проспект', 'переулок'
    ],
    'building': [
        'building', 'house', 'house_number', 'дом', 'үй', 'үйі', 'здание'
    ],
    'corpus': [
        'corpus', 'corp', 'building_block', 'корпус', 'корп', 'строение', 'стр', 'блок'
    ],
    'flat': [
        'flat', 'apartment', 'apt', 'room', 'квартира', 'кв', 'пәтер', 'комната'
    ],
    'district': [
        'district', 'area', 'region', 'район', 'аудан', 'микрорайон', 'мкр'
    ],
    'organization': [
        'organization', 'org', 'company', 'branch',
        'организация', 'предприятие', 'участок', 'участок жкх', 'домком', 'кск', 'оси'
    ]
}


def normalize_header(header: str) -> str:
    """Очищает и нормализует строку заголовка для сопоставления."""
    clean = re.sub(r'[\s_\-–—/\\.]+', ' ', str(header).lower().strip())
    clean = clean.replace('№', '').replace('#', '').strip()
    return clean


def detect_column_mapping(headers: List[str]) -> Dict[str, int]:
    """Автоматически сопоставляет заголовки файла с полями таблицы accounts."""
    mapping = {}
    normalized_headers = [normalize_header(h) for h in headers]

    for field, synonyms in COLUMN_SYNONYMS.items():
        matched_idx = None
        # Точное совпадение
        for idx, nh in enumerate(normalized_headers):
            if nh in synonyms:
                matched_idx = idx
                break
        # Частичное совпадение
        if matched_idx is None:
            for idx, nh in enumerate(normalized_headers):
                if any(syn in nh for syn in synonyms):
                    matched_idx = idx
                    break
        if matched_idx is not None:
            mapping[field] = matched_idx

    # Если 'account_number' не найден по имени, берём первую непустую колонку
    if 'account_number' not in mapping and headers:
        mapping['account_number'] = 0

    return mapping


def detect_encoding(file_path: str) -> str:
    """Определяет кодировку текстового файла (UTF-8, Windows-1251, Latin1)."""
    with open(file_path, 'rb') as f:
        raw_sample = f.read(65536)

    # Проверка BOM UTF-8
    if raw_sample.startswith(b'\xef\xbb\xbf'):
        return 'utf-8-sig'

    # Пробуем UTF-8
    try:
        raw_sample.decode('utf-8')
        return 'utf-8'
    except UnicodeDecodeError:
        pass

    # Пробуем CP1251 (стандарт 1С и Windows ЖКХ)
    try:
        raw_sample.decode('cp1251')
        return 'cp1251'
    except UnicodeDecodeError:
        pass

    return 'utf-8'


def read_csv_records(file_path: str) -> Generator[Dict[str, str], None, None]:
    """Потоково считывает строки из CSV/TSV файла с автоопределением кодировки и диалекта."""
    encoding = detect_encoding(file_path)

    with open(file_path, 'r', encoding=encoding, errors='replace') as f:
        # Читаем образец для определения разделителя
        sample = f.read(16384)
        f.seek(0)

        # Автоопределение разделителя
        delimiter = ';'
        if sample:
            sniffer = csv.Sniffer()
            try:
                dialect = sniffer.sniff(sample, delimiters=';,\t|')
                delimiter = dialect.delimiter
            except Exception:
                # Ручной подсчет популярных разделителей
                counts = {';': sample.count(';'), ',': sample.count(','), '\t': sample.count('\t'), '|': sample.count('|')}
                delimiter = max(counts, key=counts.get) if any(counts.values()) else ';'

        reader = csv.reader(f, delimiter=delimiter)
        try:
            raw_headers = next(reader, None)
        except StopIteration:
            return

        if not raw_headers:
            return

        mapping = detect_column_mapping(raw_headers)
        if 'account_number' not in mapping:
            mapping['account_number'] = 0

        for row in reader:
            if not row or not any(row):
                continue

            def extract(field: str) -> str:
                idx = mapping.get(field)
                if idx is not None and idx < len(row):
                    val = row[idx]
                    return str(val).strip() if val is not None else ''
                return ''

            acc = extract('account_number')
            # Очистка номера счёта от экспоненциального формата 1.03997e5
            if re.match(r'^\d+\.0$', acc):
                acc = acc[:-2]

            if not acc:
                continue

            addr = extract('address')
            street = extract('street')
            building = extract('building')
            corpus = extract('corpus')
            flat = extract('flat')

            # Если адрес не задан, но есть составные части — собираем
            if not addr and (street or building or flat):
                parts = []
                if street: parts.append(street)
                if building: parts.append(f"д. {building}")
                if corpus: parts.append(f"корп. {corpus}")
                if flat: parts.append(f"кв. {flat}")
                addr = ", ".join(parts)

            yield {
                'account_number': acc,
                'customer_name': extract('customer_name'),
                'address': addr,
                'street': street,
                'building': building,
                'corpus': corpus,
                'district': extract('district'),
                'organization': extract('organization')
            }


def read_excel_records(file_path: str) -> Generator[Dict[str, str], None, None]:
    """Считывает строки из файлов Excel (.xlsx или .xls)."""
    ext = os.path.splitext(file_path)[1].lower()

    if ext == '.xlsx':
        try:
            import openpyxl
            book = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet = book[book.sheetnames[0]]
            iter_rows = sheet.iter_rows(values_only=True)
            raw_headers = next(iter_rows, None)
            if not raw_headers:
                book.close()
                return

            mapping = detect_column_mapping([str(h or '') for h in raw_headers])
            if 'account_number' not in mapping:
                mapping['account_number'] = 0

            for row in iter_rows:
                if not row or not any(row):
                    continue

                def get_val(field: str) -> str:
                    idx = mapping.get(field)
                    if idx is not None and idx < len(row):
                        v = row[idx]
                        if v is None:
                            return ''
                        if isinstance(v, float) and v.is_integer():
                            return str(int(v))
                        return str(v).strip()
                    return ''

                acc = get_val('account_number')
                if not acc:
                    continue

                yield {
                    'account_number': acc,
                    'customer_name': get_val('customer_name'),
                    'address': get_val('address'),
                    'street': get_val('street'),
                    'building': get_val('building'),
                    'corpus': get_val('corpus'),
                    'district': get_val('district'),
                    'organization': get_val('organization')
                }
            book.close()
        except ImportError:
            raise RuntimeError("Для чтения .xlsx файлов установите: pip install openpyxl")

    elif ext == '.xls':
        try:
            import xlrd
            book = xlrd.open_workbook(file_path)
            sheet = book.sheet_by_index(0)
            if sheet.nrows < 1:
                return

            raw_headers = [str(sheet.cell_value(0, c)) for c in range(sheet.ncols)]
            mapping = detect_column_mapping(raw_headers)
            if 'account_number' not in mapping:
                mapping['account_number'] = 0

            for r_idx in range(1, sheet.nrows):
                def get_val(field: str) -> str:
                    idx = mapping.get(field)
                    if idx is not None and idx < sheet.ncols:
                        v = sheet.cell_value(r_idx, idx)
                        if isinstance(v, float) and v.is_integer():
                            return str(int(v))
                        return str(v).strip()
                    return ''

                acc = get_val('account_number')
                if not acc:
                    continue

                yield {
                    'account_number': acc,
                    'customer_name': get_val('customer_name'),
                    'address': get_val('address'),
                    'street': get_val('street'),
                    'building': get_val('building'),
                    'corpus': get_val('corpus'),
                    'district': get_val('district'),
                    'organization': get_val('organization')
                }
        except ImportError:
            raise RuntimeError("Для чтения .xls файлов установите: pip install xlrd")


def read_json_records(file_path: str) -> Generator[Dict[str, str], None, None]:
    """Считывает записи из JSON или JSON-Lines (.jsonl)."""
    with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
        # Проверяем, это массив JSON или построчный JSONL
        first_char = f.read(1).strip()
        f.seek(0)

        if first_char == '[':
            items = json.load(f)
            for item in items:
                if isinstance(item, dict):
                    acc = str(item.get('account_number') or item.get('account') or item.get('лс') or '').strip()
                    if acc:
                        yield {
                            'account_number': acc,
                            'customer_name': str(item.get('customer_name') or item.get('fio') or item.get('фио') or '').strip(),
                            'address': str(item.get('address') or item.get('адрес') or '').strip(),
                            'street': str(item.get('street') or item.get('улица') or '').strip(),
                            'building': str(item.get('building') or item.get('дом') or '').strip(),
                            'corpus': str(item.get('corpus') or item.get('корпус') or '').strip(),
                            'district': str(item.get('district') or item.get('район') or '').strip(),
                            'organization': str(item.get('organization') or item.get('организация') or '').strip()
                        }
        else:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    item = json.loads(line)
                    acc = str(item.get('account_number') or item.get('account') or item.get('лс') or '').strip()
                    if acc:
                        yield {
                            'account_number': acc,
                            'customer_name': str(item.get('customer_name') or item.get('fio') or item.get('фио') or '').strip(),
                            'address': str(item.get('address') or item.get('адрес') or '').strip(),
                            'street': str(item.get('street') or item.get('улица') or '').strip(),
                            'building': str(item.get('building') or item.get('дом') or '').strip(),
                            'corpus': str(item.get('corpus') or item.get('корпус') or '').strip(),
                            'district': str(item.get('district') or item.get('район') or '').strip(),
                            'organization': str(item.get('organization') or item.get('организация') or '').strip()
                        }
                except json.JSONDecodeError:
                    continue


def read_sqlite_records(file_path: str) -> Generator[Dict[str, str], None, None]:
    """Считывает лицевые счета напрямую из другого файла базы SQLite."""
    con = sqlite3.connect(file_path)
    con.row_factory = sqlite3.Row
    try:
        cur = con.cursor()
        cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='accounts'")
        if not cur.fetchone():
            raise ValueError(f"Таблица 'accounts' не найдена в SQLite файле {file_path}")

        rows = cur.execute("SELECT * FROM accounts").fetchall()
        for r in rows:
            keys = r.keys()
            acc = str(r['account_number'] if 'account_number' in keys else r['id']).strip()
            if not acc:
                continue
            yield {
                'account_number': acc,
                'customer_name': str(r['customer_name'] or '') if 'customer_name' in keys else '',
                'address': str(r['address'] or '') if 'address' in keys else '',
                'street': str(r['street'] or '') if 'street' in keys else '',
                'building': str(r['building'] or '') if 'building' in keys else '',
                'corpus': str(r['corpus'] or '') if 'corpus' in keys else '',
                'district': str(r['district'] or '') if 'district' in keys else '',
                'organization': str(r['organization'] or '') if 'organization' in keys else ''
            }
    finally:
        con.close()


def load_records_from_source(file_path: str) -> Generator[Dict[str, str], None, None]:
    """Маршрутизатор считывания записей в зависимости от расширения файла."""
    ext = os.path.splitext(file_path)[1].lower()
    if ext in ('.csv', '.tsv', '.txt'):
        return read_csv_records(file_path)
    elif ext in ('.xlsx', '.xls'):
        return read_excel_records(file_path)
    elif ext in ('.json', '.jsonl'):
        return read_json_records(file_path)
    elif ext in ('.sqlite', '.sqlite3', '.db'):
        return read_sqlite_records(file_path)
    else:
        # По умолчанию пробуем как CSV
        return read_csv_records(file_path)


def import_accounts_file(
    file_path: str,
    mode: str = 'upsert',
    batch_size: int = 2000,
    verbose: bool = True
) -> Dict[str, Any]:
    """
    Импортирует файл со счетами в целевую базу данных (PostgreSQL или SQLite).

    Режимы:
      - 'upsert': добавляет новые, обновляет существующие
      - 'insert_only': добавляет только отсутствующие
      - 'replace': полностью очищает accounts перед импортом
    """
    start_time = time.time()
    migrate_db()

    if not os.path.isfile(file_path):
        raise FileNotFoundError(f"Файл не найден: {file_path}")

    if verbose:
        db_type = "PostgreSQL" if is_postgres_configured() else "SQLite"
        print(f"🚀 [Импорт] Подключение к БД: {db_type}")
        print(f"📂 [Импорт] Файл: {file_path}")
        print(f"⚙️ [Импорт] Режим: {mode}")

    if mode == 'replace':
        if verbose:
            print("⚠ Режим 'replace': очистка существующей таблицы accounts...")
        with write_transaction() as con:
            con.execute("DELETE FROM accounts")

    inserted_count = 0
    updated_count = 0
    skipped_count = 0
    batch: List[Tuple] = []

    def flush_batch(current_batch: List[Tuple]):
        nonlocal inserted_count, updated_count
        if not current_batch:
            return

        with write_transaction() as con:
            if mode == 'insert_only':
                con.executemany('''
                    INSERT INTO accounts(
                        account_number, customer_name, address, street, building, corpus, district, organization
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(account_number) DO NOTHING
                ''', current_batch)
            else:  # upsert или replace
                con.executemany('''
                    INSERT INTO accounts(
                        account_number, customer_name, address, street, building, corpus, district, organization
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(account_number) DO UPDATE SET
                        customer_name = COALESCE(NULLIF(excluded.customer_name, ''), accounts.customer_name),
                        address = COALESCE(NULLIF(excluded.address, ''), accounts.address),
                        street = COALESCE(NULLIF(excluded.street, ''), accounts.street),
                        building = COALESCE(NULLIF(excluded.building, ''), accounts.building),
                        corpus = COALESCE(NULLIF(excluded.corpus, ''), accounts.corpus),
                        district = COALESCE(NULLIF(excluded.district, ''), accounts.district),
                        organization = COALESCE(NULLIF(excluded.organization, ''), accounts.organization)
                ''', current_batch)

        inserted_count += len(current_batch)

    record_gen = load_records_from_source(file_path)

    for item in record_gen:
        acc = item['account_number']
        if not acc:
            skipped_count += 1
            continue

        row_tuple = (
            acc,
            item['customer_name'],
            item['address'],
            item['street'],
            item['building'],
            item['corpus'],
            item['district'],
            item['organization']
        )
        batch.append(row_tuple)

        if len(batch) >= batch_size:
            flush_batch(batch)
            batch.clear()
            if verbose:
                print(f"  ...обработано счетов: {inserted_count:,}", end='\r', flush=True)

    if batch:
        flush_batch(batch)
        batch.clear()

    elapsed = round(time.time() - start_time, 2)

    # Проверяем итоговый счётчик в базе
    con_check = get_db()
    try:
        total_in_db = con_check.execute("SELECT COUNT(*) FROM accounts").fetchone()[0]
    finally:
        con_check.close()

    result = {
        'status': 'success',
        'imported': inserted_count,
        'skipped': skipped_count,
        'total_in_db': total_in_db,
        'elapsed_seconds': elapsed
    }

    if verbose:
        print(f"\n✅ [Импорт завершен] Успешно загружено: {inserted_count:,} счетов за {elapsed} сек.")
        print(f"👥 Всего лицевых счетов в базе данных: {total_in_db:,}\n")

    return result


def export_accounts_to_csv(output_path: str, verbose: bool = True) -> int:
    """Выгружает все лицевые счета из базы в CSV файл для бэкапа или переноса."""
    con = get_db()
    try:
        cur = con.execute('''
            SELECT account_number, customer_name, address, street, building, corpus, district, organization
            FROM accounts
            ORDER BY account_number ASC
        ''')
        rows = cur.fetchall()
    finally:
        con.close()

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)

    with open(output_path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f, delimiter=';')
        writer.writerow([
            'account_number', 'customer_name', 'address', 'street',
            'building', 'corpus', 'district', 'organization'
        ])
        for r in rows:
            writer.writerow([
                r['account_number'],
                r['customer_name'] or '',
                r['address'] or '',
                r['street'] or '',
                r['building'] or '',
                r['corpus'] or '',
                r['district'] or '',
                r['organization'] or ''
            ])

    if verbose:
        print(f"📦 [Экспорт завершен] Выгружено {len(rows):,} счетов в файл: {output_path}")

    return len(rows)


def main():
    parser = argparse.ArgumentParser(
        description="Импорт, экспорт и перенос базы лицевых счетов (CSV, Excel, JSON, SQLite) в PostgreSQL/SQLite."
    )
    parser.add_argument('--file', '-f', help="Путь к файлу со счетами (.csv, .xlsx, .xls, .json, .sqlite3)")
    parser.add_argument('--mode', '-m', choices=['upsert', 'insert_only', 'replace'], default='upsert',
                        help="Режим импорта: upsert (по умолчанию), insert_only, replace")
    parser.add_argument('--export', '-e', help="Путь к файлу для экспорта базы счетов в CSV (бэкап)")
    parser.add_argument('--batch-size', type=int, default=2000, help="Размер пачки для пакетной вставки (по умолчанию 2000)")

    args = parser.parse_args()

    if args.export:
        export_accounts_to_csv(args.export)
        sys.exit(0)

    if not args.file:
        parser.print_help()
        print("\nПримеры:")
        print("  python import_accounts.py --file accounts.csv")
        print("  python import_accounts.py --file registry.xlsx --mode replace")
        print("  python import_accounts.py --file old_database.sqlite3")
        print("  python import_accounts.py --export backup_accounts.csv")
        sys.exit(1)

    try:
        import_accounts_file(args.file, mode=args.mode, batch_size=args.batch_size)
    except Exception as err:
        print(f"❌ Ошибка импорта: {err}", file=sys.stderr)
        logger.error(f"[ImportAccounts] Ошибка: {err}", exc_info=True)
        sys.exit(1)


if __name__ == '__main__':
    main()
